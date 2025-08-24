from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
import json, os, httpx, time, threading, subprocess, uuid, tempfile, sys, base64, shutil
from typing import Dict, Any, List, Optional

app = FastAPI(title="Mini-Manus", version="1.5")
_JOBS: Dict[str, Dict[str, Any]] = {}
_LOCK = threading.Lock()

WORKSPACE = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "workspace"))
os.makedirs(WORKSPACE, exist_ok=True)

class RunReq(BaseModel):
    skill: str
    payload: dict = {}

def _job_finish(job_id: str, result: dict, status: str = "DONE"):
    with _LOCK:
        _JOBS[job_id]["status"] = status
        _JOBS[job_id]["result"] = result

def _load_cfg(path: str):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _tg_api(bot_token: str, method: str, **data):
    url = f"https://api.telegram.org/bot{bot_token}/{method}"
    with httpx.Client(timeout=30.0) as c:
        r = c.post(url, data=data)
        return r.json() if r.status_code == 200 else {"ok": False, "status": r.status_code, "text": r.text}

def _safe_path(p: str, allow_root: Optional[str]=None):
    root = os.path.abspath(allow_root or WORKSPACE)
    full = os.path.abspath(p if os.path.isabs(p) else os.path.join(root, p))
    if not full.startswith(root):
        raise ValueError("Caminho fora da área permitida")
    return full

@app.get("/health")
def health():
    return {"ok": True, "ts": int(time.time()), "host": os.environ.get("COMPUTERNAME",""), "platform": os.name}

@app.get("/list_skills")
def list_skills():
    return {"skills": [
        "ping_echo","notify_telegram","http_fetch","run_ps","run_py","poll_telegram",
        "file_io","excel_write","git_cli","github_file","vercel_deploy","railway_deploy",
        "loki_push","grafana_query"
    ]}

@app.post("/run")
def run(req: RunReq):
    job_id = str(uuid.uuid4())
    _JOBS[job_id] = {"status":"RUNNING", "result": None}

    def worker():
        try:
            # 1) ping
            if req.skill == "ping_echo":
                _job_finish(job_id, {"echo": req.payload.get("text","")}); return

            # 2) Telegram (saída)
            if req.skill == "notify_telegram":
                cfg = _load_cfg(req.payload.get("config_path") or "")
                tg  = (cfg.get("config",{}).get("auth",{}).get("telegram",{}) if cfg else {})
                bot_token = tg.get("bot_token")
                default_id = tg.get("chat_id_default")
                text = req.payload.get("text") or ""
                if not text: raise ValueError("payload.text obrigatório")
                dests: List[str] = []
                if isinstance(req.payload.get("to"), list): dests += [str(x) for x in req.payload["to"] if x]
                if req.payload.get("chat_id"): dests.append(str(req.payload["chat_id"]))
                if not dests and default_id: dests.append(str(default_id))
                if not bot_token or not dests: raise ValueError("bot_token ou chat_id ausente")
                ok, errs = 0, []
                for cid in dests:
                    r = _tg_api(bot_token, "sendMessage", chat_id=cid, text=text)
                    if r.get("ok"): ok += 1
                    else: errs.append({"chat_id": cid, "resp": r})
                _job_finish(job_id, {"ok": ok>0, "sent": ok, "errors": errs}); return

            # 3) HTTP
            if req.skill == "http_fetch":
                url = req.payload.get("url"); method = (req.payload.get("method") or "GET").upper()
                data = req.payload.get("data"); headers = req.payload.get("headers") or {}
                if not url: raise ValueError("payload.url obrigatório")
                with httpx.Client(timeout=45.0) as c:
                    r = c.request(method, url, data=data, headers=headers)
                try: body = r.json()
                except Exception: body = r.text[:2000]
                _job_finish(job_id, {"status": r.status_code, "headers": dict(r.headers), "body": body}); return

            # 4) PowerShell
            if req.skill == "run_ps":
                cmd = req.payload.get("cmd")
                if not cmd: raise ValueError("payload.cmd obrigatório")
                completed = subprocess.run(["powershell","-NoProfile","-Command", cmd], capture_output=True, text=True)
                _job_finish(job_id, {"rc": completed.returncode, "stdout": completed.stdout, "stderr": completed.stderr}); return

            # 5) Python
            if req.skill == "run_py":
                code = req.payload.get("code"); path = req.payload.get("path")
                if not code and not path: raise ValueError("payload.code ou payload.path obrigatório")
                run_path = path
                tmp = None
                if code:
                    tmp = tempfile.NamedTemporaryFile("w", suffix=".py", delete=False, encoding="utf-8")
                    tmp.write(code); tmp.close(); run_path = tmp.name
                completed = subprocess.run([sys.executable, run_path], capture_output=True, text=True)
                if tmp: os.unlink(tmp.name)
                _job_finish(job_id, {"rc": completed.returncode, "stdout": completed.stdout, "stderr": completed.stderr}); return

            # 6) Telegram (entrada)
            if req.skill == "poll_telegram":
                cfg = _load_cfg(req.payload.get("config_path") or "")
                tg  = (cfg.get("config",{}).get("auth",{}).get("telegram",{}) if cfg else {})
                token = tg.get("bot_token"); default_id = tg.get("chat_id_default")
                if not token: raise ValueError("bot_token ausente")
                ofs_file = os.path.join(WORKSPACE, "tg.offset")
                try: offset = int(open(ofs_file,"r").read().strip())
                except Exception: offset = 0
                updates = _tg_api(token, "getUpdates", timeout=10, offset=offset+1)
                handled = 0
                if updates.get("ok"):
                    for u in updates.get("result", []):
                        offset = max(offset, u.get("update_id", 0))
                        msg = (u.get("message") or u.get("edited_message") or {})
                        text = (msg.get("text") or "").strip()
                        chat_id = str((msg.get("chat") or {}).get("id") or default_id)
                        if not chat_id: continue
                        if text.startswith("/ping"):
                            _tg_api(token,"sendMessage", chat_id=chat_id, text="pong ✅"); handled += 1
                        elif text.startswith("/say "):
                            _tg_api(token,"sendMessage", chat_id=chat_id, text=text[5:]); handled += 1
                        elif text.startswith("/ps "):
                            cmd = text[4:]
                            r = subprocess.run(["powershell","-NoProfile","-Command", cmd], capture_output=True, text=True)
                            out = (r.stdout or r.stderr or "").strip()
                            _tg_api(token,"sendMessage", chat_id=chat_id, text=(out[:4000] or f"rc={r.returncode}")); handled += 1
                        elif text.startswith("/py "):
                            code = text[4:]
                            tmp = tempfile.NamedTemporaryFile("w", suffix=".py", delete=False, encoding="utf-8")
                            tmp.write(code); tmp.close()
                            r = subprocess.run([sys.executable, tmp.name], capture_output=True, text=True)
                            os.unlink(tmp.name)
                            out = (r.stdout or r.stderr or "").strip()
                            _tg_api(token,"sendMessage", chat_id=chat_id, text=(out[:4000] or f"rc={r.returncode}")); handled += 1
                open(ofs_file,"w").write(str(offset))
                _job_finish(job_id, {"ok": True, "handled": handled}); return

            # 7) File I/O
            if req.skill == "file_io":
                act = req.payload.get("action")
                allow_root = req.payload.get("allow_root")
                path = req.payload.get("path") or ""
                if act in ["list","ls"]:
                    folder = _safe_path(path or ".", allow_root)
                    items = []
                    for name in os.listdir(folder):
                        p = os.path.join(folder, name); st = os.stat(p)
                        items.append({"name": name, "is_dir": os.path.isdir(p), "size": st.st_size})
                    _job_finish(job_id, {"ok": True, "items": items}); return
                if act == "read":
                    p = _safe_path(path, allow_root)
                    with open(p, "rb") as f: data = f.read()
                    mode = (req.payload.get("mode") or "text")
                    if mode == "base64":
                        _job_finish(job_id, {"ok": True, "b64": base64.b64encode(data).decode()}); return
                    _job_finish(job_id, {"ok": True, "text": data.decode("utf-8", errors="replace")}); return
                if act in ["write","append"]:
                    p = _safe_path(path, allow_root)
                    os.makedirs(os.path.dirname(p), exist_ok=True)
                    text = req.payload.get("text"); b64 = req.payload.get("b64")
                    data = base64.b64decode(b64) if b64 else (text or "").encode("utf-8")
                    mode = "ab" if act=="append" else "wb"
                    with open(p, mode) as f: f.write(data)
                    _job_finish(job_id, {"ok": True, "path": p}); return
                if act == "mkdir":
                    p = _safe_path(path, allow_root); os.makedirs(p, exist_ok=True)
                    _job_finish(job_id, {"ok": True, "path": p}); return
                if act == "move":
                    src = _safe_path(req.payload.get("src"), allow_root)
                    dst = _safe_path(req.payload.get("dst"), allow_root)
                    os.makedirs(os.path.dirname(dst), exist_ok=True); shutil.move(src, dst)
                    _job_finish(job_id, {"ok": True, "src": src, "dst": dst}); return
                if act == "delete":
                    p = _safe_path(path, allow_root)
                    if os.path.isdir(p): shutil.rmtree(p)
                    elif os.path.exists(p): os.remove(p)
                    _job_finish(job_id, {"ok": True}); return
                if act == "open_folder":
                    p = _safe_path(path or ".", allow_root)
                    subprocess.Popen(["explorer.exe", p]); _job_finish(job_id, {"ok": True, "opened": p}); return
                raise ValueError("ação inválida em file_io")

            # 8) Excel
            if req.skill == "excel_write":
                try:
                    import openpyxl
                except Exception:
                    raise ValueError("Instale 'openpyxl' no venv")
                path = _safe_path(req.payload.get("path") or "planilha.xlsx", req.payload.get("allow_root"))
                sheet = req.payload.get("sheet") or "Plan1"
                data  = req.payload.get("data") or []
                os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
                if os.path.exists(path):
                    wb = openpyxl.load_workbook(path)
                else:
                    wb = openpyxl.Workbook()
                ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
                ws.delete_rows(1, ws.max_row)
                for r, row in enumerate(data, start=1):
                    for c, val in enumerate(row, start=1):
                        ws.cell(row=r, column=c, value=val)
                wb.save(path)
                _job_finish(job_id, {"ok": True, "path": path, "rows": len(data)}); return

            # 9) Git CLI
            if req.skill == "git_cli":
                act = (req.payload.get("action") or "status").lower()
                repo_dir = _safe_path(req.payload.get("dir") or "repo", req.payload.get("allow_root"))
                os.makedirs(repo_dir, exist_ok=True)
                if act == "clone":
                    url = req.payload.get("url")
                    if not url: raise ValueError("url obrigatório")
                    completed = subprocess.run(["git","clone",url,repo_dir], capture_output=True, text=True)
                elif act == "pull":
                    completed = subprocess.run(["git","-C",repo_dir,"pull"], capture_output=True, text=True)
                elif act == "status":
                    completed = subprocess.run(["git","-C",repo_dir,"status","--short","--branch"], capture_output=True, text=True)
                else:
                    raise ValueError("ação git inválida")
                _job_finish(job_id, {"rc": completed.returncode, "stdout": completed.stdout, "stderr": completed.stderr}); return

            # 10) GitHub API
            if req.skill == "github_file":
                cfg = _load_cfg(req.payload.get("config_path") or "")
                token = ((cfg.get("config",{}).get("auth",{}).get("github",{}) or {}).get("token"))
                if not token: raise ValueError("GitHub token ausente")
                owner = req.payload.get("owner"); repo = req.payload.get("repo"); path = req.payload.get("path")
                content_text = req.payload.get("content") or ""
                message = req.payload.get("message") or "update via API"
                if not (owner and repo and path): raise ValueError("owner/repo/path obrigatórios")
                api = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}"
                headers = {"Authorization": f"Bearer {token}", "Accept":"application/vnd.github+json"}
                sha = None
                with httpx.Client(timeout=45.0) as c:
                    r = c.get(api, headers=headers)
                    if r.status_code == 200: sha = r.json().get("sha")
                    b64 = base64.b64encode(content_text.encode("utf-8")).decode()
                    payload = {"message": message, "content": b64}
                    if sha: payload["sha"] = sha
                    r2 = c.put(api, headers=headers, json=payload)
                    _job_finish(job_id, {"status": r2.status_code, "body": r2.json()}); return

            # 11) Vercel CLI
            if req.skill == "vercel_deploy":
                cfg = _load_cfg(req.payload.get("config_path") or "")
                token = ((cfg.get("config",{}).get("auth",{}).get("vercel",{}) or {}).get("token"))
                proj_dir = _safe_path(req.payload.get("dir") or WORKSPACE, req.payload.get("allow_root"))
                if not token: raise ValueError("vercel token ausente")
                env = os.environ.copy(); env["VERCEL_TOKEN"] = token
                args = ["vercel","--prod","-y"]
                completed = subprocess.run(args, cwd=proj_dir, env=env, capture_output=True, text=True)
                _job_finish(job_id, {"rc": completed.returncode, "stdout": completed.stdout, "stderr": completed.stderr}); return

            # 12) Railway CLI (com RAILWAY_TOKEN)
            if req.skill == "railway_deploy":
                cfg = _load_cfg(req.payload.get("config_path") or "")
                token = ((cfg.get("config",{}).get("auth",{}).get("railway",{}) or {}).get("token"))
                proj_dir = _safe_path(req.payload.get("dir") or WORKSPACE, req.payload.get("allow_root"))
                if not token: raise ValueError("railway token ausente")
                env = os.environ.copy()
                env["RAILWAY_TOKEN"] = token
                args = ["railway","up"]
                service = (req.payload.get("service") or "").strip()
                if service: args += ["--service", service]
                completed = subprocess.run(args, cwd=proj_dir, env=env, capture_output=True, text=True)
                _job_finish(job_id, {"rc": completed.returncode, "stdout": completed.stdout, "stderr": completed.stderr}); return

            # 13) Loki push
            if req.skill == "loki_push":
                cfg = _load_cfg(req.payload.get("config_path") or "")
                loki_url = ((cfg.get("config",{}).get("obs",{}).get("loki",{}) or {}).get("url"))
                if not loki_url: raise ValueError("loki url ausente")
                labels = req.payload.get("labels") or {"job":"manus","host":os.environ.get("COMPUTERNAME","")}
                msg = req.payload.get("text") or "ping"
                now_ns = int(time.time() * 1e9)
                payload = {"streams":[{"stream":labels, "values":[[str(now_ns), msg]]}]}
                with httpx.Client(timeout=30.0) as c:
                    r = c.post(loki_url, json=payload)
                _job_finish(job_id, {"status": r.status_code, "text": r.text[:2000]}); return

            # 14) Grafana query
            if req.skill == "grafana_query":
                cfg = _load_cfg(req.payload.get("config_path") or "")
                g = (cfg.get("config",{}).get("obs",{}).get("grafana",{}) or {})
                url = g.get("url"); key = g.get("api_key")
                if not (url and key): raise ValueError("grafana url/api_key ausentes")
                endpoint = req.payload.get("endpoint") or "/api/search"
                params = req.payload.get("params") or {"query":""}
                with httpx.Client(timeout=45.0, headers={"Authorization": f"Bearer {key}"}) as c:
                    r = c.get(url.rstrip("/") + endpoint, params=params)
                try: body = r.json()
                except Exception: body = r.text[:2000]
                _job_finish(job_id, {"status": r.status_code, "body": body}); return

            raise ValueError(f"skill desconhecida: {req.skill}")
        except Exception as e:
            _job_finish(job_id, {"error": str(e)}, status="ERROR")

    threading.Thread(target=worker, daemon=True).start()
    return {"job_id": job_id}

@app.get("/status/{job_id}")
def status(job_id: str):
    if job_id not in _JOBS: raise HTTPException(status_code=404, detail="job_id não encontrado")
    return _JOBS[job_id]
